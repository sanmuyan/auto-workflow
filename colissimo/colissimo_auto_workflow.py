# -*- coding: utf-8 -*-
import logging
import re
import time
import hashlib
from PIL import Image, ImageGrab
import pytesseract
import signal
import sys

from openpyxl.reader.excel import load_workbook
from openpyxl.drawing.image import Image as openpyxl_image

logging.basicConfig(format="%(asctime)s %(levelname)s:%(message)s", datefmt="%Y-%m-%d %H:%M:%S", level=logging.INFO)


def save_parcel_data(parcel_number):
    try:
        wb = load_workbook(table_file)
        ws = wb.active
        last_row_number = ws.max_row
        last_a_row_data = ws.cell(row=last_row_number, column=1).value
        try:
            last_a_row_data = int(last_a_row_data)
            new_a_row_data = last_a_row_data + 1
        except ValueError as e:
            new_a_row_data = 1
            last_row_number = 1
        new_row_number = last_row_number + 1
        logging.info(f'新增行={new_row_number}')
        new_c_row_data = parcel_number
        new_d_row_data = openpyxl_image(screenshot_image_file)
        ws[f'A{new_row_number}'] = new_a_row_data
        ws[f'C{new_row_number}'] = new_c_row_data
        ws.row_dimensions[new_row_number].height = 50
        d_row_height = 50
        d_row_width = int(new_d_row_data.width * (d_row_height / new_d_row_data.height))
        new_d_row_data.anchor = f'D{new_row_number}'
        new_d_row_data.height = d_row_height
        new_d_row_data.width = d_row_width
        ws.add_image(new_d_row_data)
        wb.save(table_file)
        wb.close()
    except Exception as e:
        logging.error(f'打开表格异常，错误={e}')


def get_parcel_number(clipboard_image):
    ocr_text = pytesseract.image_to_string(clipboard_image, lang="eng")
    ocr_text = ocr_text.strip()
    parcel_number = None
    for line in ocr_text.split('\n'):
        if 'Parcel No.' in line:
            parcel_number = None
            m = re.search(r'6.*?\b', line)
            if m:
                parcel_number = m.group()
                if len(parcel_number) != 13:
                    logging.warning(f'忽略非法运单={parcel_number}')
                if parcel_number[:2] == '64':
                    parcel_number = '6A' + parcel_number[2:]
            else:
                logging.warning(f'未能截取有效运单号={line}')
            break
    if parcel_number:
        if parcel_number not in parcel_list:
            logging.info(f'保存运单={parcel_number}')
            clipboard_image.save(screenshot_image_file)
            try:
                save_parcel_data(parcel_number)
                parcel_list.append(parcel_number)
                logging.info(f'保存成功={parcel_number}')
            except Exception as e:
                logging.error(f'保存表格异常，错误={e}')
        else:
            logging.warning(f'忽略重复运单={parcel_number}')


def watch_clipboard_images():
    hash_list = []
    while True:
        try:
            clipboard_image = ImageGrab.grabclipboard()
            if clipboard_image:
                im_bytes = clipboard_image.tobytes()
                im_hash = hashlib.sha256(im_bytes).hexdigest()
                if im_hash not in hash_list:
                    hash_list.append(im_hash)
                    clipboard_image = Image.frombytes(clipboard_image.mode, clipboard_image.size, im_bytes)
                    get_parcel_number(clipboard_image)
                else:
                    # logging.warning(f'忽略重复截图')
                    pass
        except Exception as e:
            logging.error(f'获取剪贴板图片异常，错误={e}')
        time.sleep(0.5)


def main():
    logging.info("开始监听运单截图...")
    try:
        wb = load_workbook(table_file)
        ws = wb.active
        for i in range(2, ws.max_row + 1):
            parcel_list.append(ws.cell(row=i, column=3).value)
        wb.close()
    except Exception as _e:
        logging.error(f'打开表格异常，错误={_e}')
        sys.exit(0)
    watch_clipboard_images()


def exit_handler(signum, frame):
    logging.warning(f'shutdown bot by os signal {signum}')
    sys.exit(0)


if __name__ == '__main__':
    signal.signal(signal.SIGINT, exit_handler)
    table_file = './files/colissimo_auto.xlsx'
    screenshot_image_file = './files/screenshot_image.png'
    parcel_list = []
    main()
