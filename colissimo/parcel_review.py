# -*- coding: utf-8 -*-

from openpyxl.reader.excel import load_workbook

if __name__ == '__main__':
    table_file = './files/colissimo_auto.xlsx'
    require_parcel_file = './files/require_parcel.txt'
    parcel_list = []
    require_parcel_list = []

    with open(require_parcel_file, 'r') as f:
        for line in f.readlines():
            require_parcel_list.append(line.strip())
    wb = load_workbook(table_file)
    ws = wb.active
    for i in range(2, ws.max_row + 1):
        parcel_list.append(ws.cell(row=i, column=3).value)
    wb.close()
    print('漏掉的运单：')
    for parcel in require_parcel_list:
        if parcel not in parcel_list:
            print(parcel)

    print('多余的运单：')
    for parcel in parcel_list:
        if parcel not in require_parcel_list:
            print(parcel)
